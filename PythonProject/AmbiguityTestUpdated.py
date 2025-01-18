# Imports
import pandas as pd
import openpyxl
from openai import OpenAI
from docx import Document
from openpyxl.styles import Font, PatternFill, Alignment


# Module-level constants
CHAT_GPT_VERSION = "gpt-3.5-turbo-16k"  #"gpt-4o"
INPUT_DATA_EXCEL_FILE = "SeminarData.xlsx"
OUTPUT_FILE = "test_output " + CHAT_GPT_VERSION + ".xlsx"
HEADERS = ["Story File Name", "Result Text", "Sentence", "Received", "Sentence Type", "Is Correct"]
API_KEY = "" # Complete your key here

def main():
    # Load the Excel file
    data = load_and_clean_data(INPUT_DATA_EXCEL_FILE)

    for index, row in data.iterrows():
        print(row['story_file_name'], row['sentence'])

    results = []
    for i in range(4):

        # Process each row in the Excel file
        for index, row in data.iterrows ():
            # Read the story file name
            story_file_name = row['story_file_name']
            sentence = row['sentence']
            sentence_type = row['sentence_type']

            # Read and construct the prompt from the Word file
            try:
                story_doc = Document(story_file_name)
                story_text = "\n".join(paragraph.text for paragraph in story_doc.paragraphs)
            except Exception as e:
                print(f"Error reading Word file {story_file_name}: {e}")
                continue

            # test a story given sentence
            result_text = call_openai_api (story_text, sentence, API_KEY)
            logical_answer = result_text.split('.')[0].strip()

            is_answer_correct = 'correct' if logical_answer == 'True' else 'incorrect'

            # Append output results to Excel worksheet
            results.append([
                story_file_name,
                result_text,
                sentence,
                logical_answer,
                sentence_type,
                is_answer_correct
            ])

    write_output (results)

def load_and_clean_data(file_path):
    """Load Excel data and clean column names."""
    data = pd.read_excel(file_path)
    data.columns = data.columns.str.lower()
    print("Cleaned column names:", data.columns.tolist())
    return data

# Function definition for call_openai_api
def call_openai_api(story_text, sentence, key):
    """Function to make the API call."""
    intro = "Below is a story followed by an ambiguous sentence. Based on the context of the story, decide if the sentence is correct or not. Start your response with 'True.' or 'False. and explain your reasoning."
    prompt = f"{intro}\n\n{story_text}\n\n{sentence} (test Case)"

    try:
        client = OpenAI( api_key=key)
        response = client.chat.completions.create(
            model = CHAT_GPT_VERSION,
            messages=[
                {"role": "system", "content": "You are a helpful assistant who answers common-sense reasoning questions."},
                {"role": "user", "content": prompt},
            ]
        )
        return response.choices[0].message.content
    except Exception as e:
        print(f"Error calling OpenAI API: {e}")
        return f"Error: {e}"

def write_output (results):
    # Create a new workbook and select the active worksheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

     # Write column headers
    worksheet.append(HEADERS)

    for result in results:
        if result:  # Skip None values
            worksheet.append(result)

    # Get the last column index
    last_column = worksheet.max_column
    yellow_bg_color = "FFFF00"
    red_font_color="FF0000"

    # Set a title for the worksheet
    worksheet.title = "Results"

    worksheet.column_dimensions['A'].width = 22  # Column A for Story File Name
    worksheet.column_dimensions['B'].width = 120  # Column B for Result Text
    worksheet.column_dimensions['C'].width = 38  # Column C for sentence
    worksheet.column_dimensions['E'].width = 10  # Column D for Received
    worksheet.column_dimensions['F'].width = 10  # Column E for Sentence Type
    worksheet.column_dimensions['G'].width = 10  # Column F for Is Answer Correct

    # Iterate over the cells in the last column (excluding the header)
    for row in worksheet.iter_rows(min_row=2, min_col=last_column, max_col=last_column):
        for cell in row:
            if cell.value == "incorrect":
                # Apply red font color for "incorrect"
                cell.font = Font(color=red_font_color)  # Red font

    # Change alignment for the second cell (column B) in each row
    for row in worksheet.iter_rows():  # Adjust range as needed
        second_cell = row[1]  # The second cell in the row (index 1 corresponds to column B)
        second_cell.alignment = Alignment(horizontal='left', vertical='center', wrapText=True)

    header_fill = PatternFill(start_color=yellow_bg_color, end_color=yellow_bg_color, fill_type="solid")

    # Apply the background color to the first row
    for cell in worksheet[1]:  # Access the first row
        cell.fill = header_fill

    # Save the output Excel file
    workbook.save(OUTPUT_FILE)


if __name__ == "__main__":
    main()






