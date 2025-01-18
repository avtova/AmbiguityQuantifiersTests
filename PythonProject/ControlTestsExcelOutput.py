# Imports
import pandas as pd
import openpyxl
from docx import Document
from openpyxl.styles import Font, PatternFill, Alignment
from openai import OpenAI

# Module-level constants
CHAT_GPT_VERSION = "gpt-3.5-turbo-16k" #"gpt-4o"
INPUT_DATA_EXCEL_FILE = "ControlDataOnly.xlsx"
OUTPUT_FILE = "control_output " + CHAT_GPT_VERSION + ".xlsx"
HEADERS = ["Story File Name", "Result Text", "Expected", "Sent to GPT", "Received", "Is Correct"]
API_KEY = "" # Complete your key here

def main():

    # Load the Excel file
    data = load_and_clean_data (INPUT_DATA_EXCEL_FILE)

    for index, row in data.iterrows():
        print(row['story_file_name'], row['option_a'], row['option_b'])

    current_option = 'option_b'
    # Process each row from input file

    results =[]
    for i in range(6):

        # use option_a sentence and option_b sentence interchangeably for all the sentences
        current_option = 'option_a' if current_option == 'option_b' else 'option_b'

        for index, row in data.iterrows ():

            # Read the story file name
            story_file_name = row['story_file_name']
            sentence = row[current_option]
            expected_answer = row['expected_answer']

            # Read and construct the prompt from the Word file
            try:
                story_doc = Document(story_file_name)
                story_text = "\n".join(paragraph.text for paragraph in story_doc.paragraphs)
            except Exception as e:
                print(f"Error reading Word file {story_file_name}: {e}")
                continue

            result_text = call_openai_api(story_text, sentence, API_KEY)

            logical_answer = result_text.split('.')[0].strip()

            if expected_answer == current_option:
                is_answer_correct = 'correct' if logical_answer == 'True' else 'incorrect'
            else:
                is_answer_correct = 'correct' if logical_answer == 'False' else 'incorrect'

            # Append output results to Excel worksheet
            results.append([
                story_file_name,
                result_text,
                expected_answer,
                current_option,
                logical_answer,
                is_answer_correct
            ])

    # Save output results to Excel file
    write_output(results)

def load_and_clean_data(file_path):
    """Load Excel data and clean column names."""
    data = pd.read_excel(file_path)
    data.columns = data.columns.str.lower()
    print("Cleaned column names:", data.columns.tolist())
    return data

# Function definition for call_openai_api
def call_openai_api(story_text, sentence, key):
    """Function to make the API call."""
    intro = "Below is a story followed by a sentence. Based on the context of the story, decide if the sentence is correct or not. Start your response with 'True.' or 'False.' and explain your reasoning."
    prompt = f"{intro}\n\n{story_text}\n\n{sentence} "

    try:
        client = OpenAI(api_key=key)
        response = client.chat.completions.create(
            model = CHAT_GPT_VERSION,
            messages=[
                {"role": "system", "content": "You are a helpful assistant who answers common-sense reasoning questions."},
                {"role": "user", "content": prompt},
            ]
        )
        return response.choices[0].message.content
    except Exception as e:
        print(f"Error calling OpenAI API: {e}")
        return f"Error: {e}"

def write_output (results):
    # Create a new workbook and select the active worksheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

     # Write column headers
    worksheet.append(HEADERS)

    for result in results:
        if result:  # Skip None values
            worksheet.append(result)

    # Get the last column index
    last_column = worksheet.max_column
    yellow_bg_color = "FFFF00"
    red_font_color="FF0000"

    # Set a title for the worksheet
    worksheet.title = "Results"

    worksheet.column_dimensions['A'].width = 25  # Column A for Story File Name
    worksheet.column_dimensions['B'].width = 120  # Column B for Result Text
    worksheet.column_dimensions['C'].width = 10  # Column C for Expected Answer
    worksheet.column_dimensions['D'].width = 10  # Column D for Sent to GPT
    worksheet.column_dimensions['E'].width = 10  # Column E for Received
    worksheet.column_dimensions['G'].width = 10  # Column G for Is Answer Correct

    # Iterate over the cells in the last column (excluding the header)
    for row in worksheet.iter_rows(min_row=2, min_col=last_column, max_col=last_column):
        for cell in row:
            if cell.value == "incorrect":
                # Apply red font color for "incorrect"
                cell.font = Font(color=red_font_color)  # Red font

    # Wrap text for column Result Text
    for row in worksheet.iter_rows(min_row=2, min_col=2, max_col=2):  # Column B
        for cell in row:
            cell.alignment = Alignment(wrap_text=True)

    header_fill = PatternFill(start_color=yellow_bg_color, end_color=yellow_bg_color, fill_type="solid")

    # Apply the background color to the first row
    for cell in worksheet[1]:  # Access the first row
        cell.fill = header_fill

    # Save the output Excel file
    workbook.save(OUTPUT_FILE)

if __name__ == "__main__":
    main()







